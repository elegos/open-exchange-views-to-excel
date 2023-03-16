import os
import re
import xml.etree.ElementTree as ET

from argparse import ArgumentParser
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side

@dataclass
class ViewElement:
    id: str
    name: str
    documentation: str


def nstag(ns: str, tag: str) -> str:
    return '{' + ns + '}' + tag

def parse_node(node: ET.Element, view_elements: List[ViewElement], elements: ET.Element, xsi: str) -> None:
    id = node.attrib['elementRef']
    element_node = elements.find(f"{nstag(ns, 'element')}[@identifier='{id}']")
    name = element_node.find(nstag(ns, 'name')).text

    documentation = ''
    element_documentation = element_node.find(nstag(ns, 'documentation'))
    if element_documentation is not None:
        documentation = element_documentation.text
    
    if element_node.attrib[nstag(xsi, 'type')] != 'Grouping':
        view_elements.append(ViewElement(id, name, documentation))

    for node in node.findall(f"{nstag(ns, 'node')}[@{nstag(xsi, 'type')}='Element']"):
        parse_node(node, view_elements, elements, xsi)
    

def parse_view(ns: str, view: ET.Element, elements: ET.Element) -> Tuple[str, List[ViewElement]]:
    view_name = view.find(nstag(ns, 'name')).text
    xsi = 'http://www.w3.org/2001/XMLSchema-instance'

    view_elements: List[ViewElement] = []
    for node in view.findall(f"{nstag(ns, 'node')}[@{nstag(xsi, 'type')}='Element']"):
        parse_node(node, view_elements, elements, xsi)

    return (view_name, view_elements) 


args = ArgumentParser()
args.add_argument('--input', '-i', required=False)
args.add_argument('--output-dir', '-o', required=False)
args.add_argument('--view-filter', '-f', required=False)
args = args.parse_args()

output_path = Path(__file__).parent.joinpath('output')
if args.output_dir:
    output_path = Path(args.output_dir)

os.makedirs(output_path, exist_ok=True)

files = Path(__file__).parent.joinpath('input').glob('*.xml')
if args.input:
    files = [Path(args.input)]

for file_path in files:
    tree = ET.parse(file_path)
    root = tree.getroot()
    ns = re.match(r'^{([^}]+)}.*$', root.tag)[1]
    views_elem = root.find(nstag(ns, 'views'))
    diagrams_elem = views_elem.find(nstag(ns, 'diagrams'))
    elements_elem = root.find(nstag(ns, 'elements'))


    workbook = Workbook()
    views: List[Tuple[str, List[ViewElement]]] = []
    for view in diagrams_elem.findall(nstag(ns, 'view')):
        views.append(parse_view(ns, view, elements_elem))

    views.sort(key=lambda elem: elem[0])
    
    header_fill = PatternFill(start_color='90EE90', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def passes_filter(name: str) -> bool:
        if not args.view_filter:
            return True

        return re.match(args.view_filter, name) is not None

    for view in views:
        name, elements = view

        if not passes_filter(name):
            continue

        elements.sort(key=lambda el: el.name)
        sheet = workbook.create_sheet(name[0:31])

        sheet['A1'] = 'Nome entità'
        sheet['B1'] = 'Descrizione entità'

        sheet[f'A1'].fill = header_fill
        sheet[f'A1'].border = thin_border
        sheet[f'B1'].fill = header_fill
        sheet[f'B1'].border = thin_border

        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['B'].width = 80

        index = 2
        for element in elements:
            sheet[f'A{index}'] = element.name
            sheet[f'A{index}'].alignment = Alignment(horizontal='left', vertical='top')
            sheet[f'A{index}'].border = thin_border
            sheet[f'B{index}'] = element.documentation
            sheet[f'B{index}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            sheet[f'B{index}'].border = thin_border
            index += 1

    del workbook['Sheet']
    workbook.save(output_path.joinpath(file_path.stem + '.xlsx'))